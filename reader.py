"""
reader.py — Legge e struttura i dati dalla pagina C.E. scaricando da Google Sheets in RAM.
"""
import openpyxl
from openpyxl.utils import column_index_from_string as cidx
import requests
from io import BytesIO

# L'ID del tuo Google Sheet
SHEET_ID = "1HlG6IDmzkwmpZ5LdfiQB5TM5Ma7y5tVzbwTC65u5K5k"

# Link di esportazione nativo in Excel (funziona perché hai impostato "Chiunque abbia il link")
EXCEL_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
SHEET_NAME = "C.E."

# Mapping colonna valore -> nome mese
MONTH_MAP = [
    ("D",  "Gennaio"),
    ("H",  "Febbraio"),
    ("L",  "Marzo"),
    ("P",  "Aprile"),
    ("T",  "Maggio"),
    ("X",  "Giugno"),
    ("AB", "Luglio"),
    ("AF", "Agosto"),
    ("AJ", "Settembre"),
    ("AN", "Ottobre"),
    ("AR", "Novembre"),
    ("AV", "Dicembre"),
]

# Per ogni mese: colonne var, var%, peso%
MONTH_EXTRA = {
    "D":  ("E",  "F",  "G"),
    "H":  ("I",  "J",  "K"),
    "L":  ("M",  "N",  "O"),
    "P":  ("Q",  "R",  "S"),
    "T":  ("U",  "V",  "W"),
    "X":  ("Y",  "Z",  "AA"),
    "AB": ("AC", "AD", "AE"),
    "AF": ("AG", "AH", "AI"),
    "AJ": ("AK", "AL", "AM"),
    "AN": ("AO", "AP", "AQ"),
    "AR": ("AS", "AT", "AU"),
    "AV": ("AW", "AX", "AY"),
}

# Colonne income (rows 21-24): valore mese in stessa colonna principale
INCOME_ROWS = {
    21: "Entrate Lorde Stipendio",
    22: "Altre Entrate",
    23: "Entrate Lorde",
    24: "Entrate Nette",
}
PREV_YEAR_ROWS = {
    27: "Entrate Lorde Stipendio",
    28: "Altre Entrate",
    29: "Entrate Lorde",
    30: "Entrate Nette",
}


def _safe(v):
    """Restituisce None per errori Excel o estrae i numeri dalle valute."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if any(x in s for x in ["DIV","REF","VALUE","N/A","#"]):
            return None
        clean = s.replace("€","").replace("$","").replace(" ","").replace("\xa0","").replace("−","-")
        if "." in clean and "," in clean:
            clean = clean.replace(".","").replace(",",".")
        elif "," in clean:
            clean = clean.replace(",",".")
        try:
            return float(clean)
        except ValueError:
            return None
    return v


def load_data(path=None):
    """
    Carica tutti i dati dalla pagina C.E. scaricando il file XLSX direttamente in RAM.
    """
    # Usiamo User-Agent per simulare un browser e aggirare i blocchi automatici di Google
    resp = requests.get(EXCEL_URL, timeout=20, headers={"User-Agent": "Mozilla/5.0"})
    
    if resp.status_code != 200:
        raise Exception(f"Impossibile scaricare il file Excel (Status Code {resp.status_code}).")
    
    # Se il server ci restituisce una pagina web di blocco invece di un file Excel, ce ne accorgiamo
    if b"<html" in resp.content.lower() or b"<body" in resp.content.lower():
        raise Exception("Blocco di sicurezza di Google rilevato. Verifica che il file sia condiviso con 'Chiunque abbia il link'.")

    # Carichiamo il workbook in memoria RAM senza salvare su disco locale
    wb = openpyxl.load_workbook(filename=BytesIO(resp.content), data_only=True)
    
    if SHEET_NAME not in wb.sheetnames:
        raise Exception(f"Il foglio '{SHEET_NAME}' non esiste nel file scaricato.")
    ws = wb[SHEET_NAME]

    # ── Asset names e valori di partenza (col A, B, righe 9-19) ──────────────
    assets = []
    for row in range(9, 20):
        name = ws.cell(row=row, column=1).value
        start = _safe(ws.cell(row=row, column=2).value)
        if name and isinstance(name, str) and name.strip():
            assets.append({"name": name.strip(), "start": start or 0})

    start_total = _safe(ws.cell(row=8, column=2).value) or 0

    # ── Obiettivi annuali (col AZ = 52) ──────────────────────────────────────
    obj_total = _safe(ws.cell(row=8, column=52).value)
    objectives = {}
    for i, a in enumerate(assets):
        obj = _safe(ws.cell(row=9 + i, column=52).value)
        objectives[a["name"]] = obj

    # ── Dati mensili patrimonio ───────────────────────────────────────────────
    monthly_patrimonio = {}

    for col_letter, month_name in MONTH_MAP:
        col = cidx(col_letter)
        var_col, varpct_col, weight_col = MONTH_EXTRA[col_letter]

        total_val = _safe(ws.cell(row=8, column=col).value)
        filled = total_val is not None and total_val > 0

        asset_data = []
        for i, a in enumerate(assets):
            row = 9 + i
            v   = _safe(ws.cell(row=row, column=col).value)
            var = _safe(ws.cell(row=row, column=cidx(var_col)).value)
            vp  = _safe(ws.cell(row=row, column=cidx(varpct_col)).value)
            w   = _safe(ws.cell(row=row, column=cidx(weight_col)).value)
            asset_data.append({
                "name":    a["name"],
                "value":   v,
                "var_eur": var,
                "var_pct": vp,
                "weight":  w,
            })

        monthly_patrimonio[month_name] = {
            "totale":  total_val,
            "assets":  asset_data,
            "filled":  filled,
        }

    # ── Dati mensili entrate (righe 21-24) ───────────────────────────────────
    monthly_income = {}

    for col_letter, month_name in MONTH_MAP:
        col = cidx(col_letter)
        stipendio = _safe(ws.cell(row=21, column=col).value)
        altre     = _safe(ws.cell(row=22, column=col).value)
        lorde     = _safe(ws.cell(row=23, column=col).value)
        nette     = _safe(ws.cell(row=24, column=col).value)

        # Se lorde è None ma stipendio e altre ci sono, calcoliamo
        if lorde is None and stipendio is not None and altre is not None:
            lorde = stipendio + altre

        monthly_income[month_name] = {
            "stipendio": stipendio,
            "altre":     altre,
            "lorde":     lorde,
            "nette":     nette,
            "filled":    any(x is not None for x in [stipendio, altre, lorde, nette]),
        }

    # ── Totali e medie entrate (BB-BD, righe 21-24) ──────────────────────────
    income_summary = {}
    for row, label in INCOME_ROWS.items():
        avg   = _safe(ws.cell(row=row, column=54).value)  # BB = 54
        total = _safe(ws.cell(row=row, column=55).value)  # BC = 55
        yoy   = _safe(ws.cell(row=row, column=56).value)  # BD = 56
        income_summary[label] = {"avg": avg, "total": total, "yoy": yoy}

    # ── Anno precedente (righe 27-30) ────────────────────────────────────────
    prev_year = {}
    for row, label in PREV_YEAR_ROWS.items():
        v = _safe(ws.cell(row=row, column=2).value)
        prev_year[label] = v

    # ── CAGR target (col AZ row 20) ──────────────────────────────────────────
    cagr_ytd = _safe(ws.cell(row=20, column=52).value)

    wb.close()

    return {
        "excel_path":          SHEET_ID,
        "start_total":         start_total,
        "assets":              assets,
        "objectives":          objectives,
        "obj_total":           obj_total,
        "monthly_patrimonio":  monthly_patrimonio,
        "monthly_income":      monthly_income,
        "income_summary":      income_summary,
        "prev_year":           prev_year,
        "cagr_ytd":            cagr_ytd,
        "months_order":        [m for _, m in MONTH_MAP],
    }


def get_filled_months(data):
    """Restituisce i mesi con dati inseriti, in ordine."""
    return [m for m in data["months_order"] if data["monthly_patrimonio"][m]["filled"]]
